import sqlite3
from openpyxl import load_workbook
from tkinter import filedialog as fd, messagebox as mb
from os import system


def rasp_excel_reader():
    conn = sqlite3.connect('BD.db')
    cursor = conn.cursor()

    system('taskkill /IM EXCEL.EXE /F')

    timetable = list()  # расписание всех дисциплин

    loc_rasp = fd.askopenfilename()

    wb = load_workbook(loc_rasp)
    sheets_list = wb.sheetnames

    # проход по всем листам excel-файла
    for month in sheets_list:
        num_sheet_rasp = wb[month]
        max_row = num_sheet_rasp.max_row

        weeks = dict()  # словарь номеров недель (ключ) и их расположения по строкам в документе (нужно для парсинга пар)
        row_week = dict()  # словарь соответствия номера недели и максимального кол-ва строк в ней

        dis = []  # параметры дисциплины (id, неделя, дата, название, тип, и т.д.)
        id_dis = ''  # уникальный идентификатор пары
        lesson_num = ''  # номер пары (1-я, 2-я, и т.д.)

        # подсчет количества учебных недель в месяце
        for num_week in range(6, max_row + 1):
            if num_sheet_rasp.cell(row=num_week, column=1).value is not None:
                weeks[num_sheet_rasp.cell(row=num_week, column=1).value] = num_week

        # вычисление максимального кол-ва строк для каждой недели
        for week in range(min(weeks.keys()), max(weeks.keys())+1):
            if week in range(1, 5):
                num_rows_in_week = weeks[week + 1] - weeks[week]
            elif month is 'Сентябрь' and week is 5:
                num_rows_in_week = max_row - weeks[week]

            row_week[week] = num_rows_in_week

        # Парсим расписание
        for num_week in weeks:
            for day in range(2, 9):
                for row in range(weeks[num_week], weeks[num_week] + row_week[num_week]+1):
                    if type(num_sheet_rasp.cell(row=row, column=day).value) is int:
                        date = str(num_sheet_rasp.cell(row=row, column=day).value)
                    else:
                        if num_sheet_rasp.cell(row=row, column=day).value is not None:
                            dis.append(num_sheet_rasp.cell(row=row, column=day).value)
                        else:
                            if len(dis) != 0:
                                # Определение номера пары (1-я, 2-я, и т.д.)
                                if dis[3] == '09.00-10.35':
                                    lesson_num = '1'
                                elif dis[3] == '10.45-12.20':
                                    lesson_num = '2'
                                elif dis[3] == '13.00-14.35':
                                    lesson_num = '3'
                                elif dis[3] == '14.45-16.20':
                                    lesson_num = '4'
                                elif dis[3] == '16.30-18.05':
                                    lesson_num = '5'
                                elif dis[3] == '18.15-19.50':
                                    lesson_num = '6'
                                elif dis[3] == '20.00-21.35':
                                    lesson_num = '7'

                                id_dis = int(str(num_week) + str(date) + lesson_num)
                                lesson_num = ''

                                dis[4].format('\\xa0', '')

                                dis.insert(0, id_dis)
                                dis.insert(1, num_week)

                                if month == 'Сентябрь':
                                    date_month = str(date) + '.09'
                                    dis.insert(2, date_month)
                                elif month == 'Октябрь':
                                    date_month = str(date) + '.10'
                                    dis.insert(2, date_month)
                                elif month == 'Ноябрь':
                                    date_month = str(date) + '.11'
                                    dis.insert(2, date_month)
                                elif month == 'Декабрь':
                                    date_month = str(date) + '.12'
                                    dis.insert(2, date_month)
                                elif month == 'Февраль':
                                    date_month = str(date) + '.02'
                                    dis.insert(2, date_month)
                                elif month == 'Март':
                                    date_month = str(date) + '.03'
                                    dis.insert(2, date_month)
                                elif month == 'Апрель':
                                    date_month = str(date) + '.04'
                                    dis.insert(2, date_month)
                                elif month == 'Май':
                                    date_month = str(date) + '.05'
                                    dis.insert(2, date_month)
                                elif month == 'Июнь':
                                    date_month = str(date) + '.06'
                                    dis.insert(2, date_month)

                                timetable.append(dis)

                            id_dis = ''
                            dis = []

    # загрузка данных в БД
    try:
        for i in range(0, len(timetable)):
            cursor.execute("""INSERT INTO ДИСЦИПЛИНА VALUES (?,?,?,?,?,?,?,?)""", timetable[i])
        conn.commit()
    except sqlite3.IntegrityError:
        mb.showerror('Внимание', 'Такое расписание уже есть в базе')

    mb.askokcancel('Загрузка расписания', 'Загрузка расписания завершена.')

    system('taskkill /IM EXCEL.EXE /F')
