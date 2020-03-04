from pathlib import Path
import sqlite3
from openpyxl import load_workbook
from tkinter import filedialog as fd, messagebox as mb
from os import system


def group_excel_reader():
    conn = sqlite3.connect('BD.db')
    cursor = conn.cursor()

    system('taskkill /IM EXCEL.EXE /F')

    loc_group = fd.askopenfilename()

    wb = load_workbook(loc_group)

    num_sheet = wb['Лист1']
    max_row = num_sheet.max_row

    group_num = Path(loc_group).stem

    global students_list

    for row in range(1, max_row + 1):
        if num_sheet.cell(row=row, column=2) is not None:
            student = num_sheet.cell(row=row, column=2).value
            a = student.split(' ')

            # проверка на наличие отчества у студента
            if len(a) == 2:
                a.append('0')

            a.append(group_num)

            # Проверка на наличие записи в БД
            if cursor.execute("""SELECT * FROM СТУДЕНТ WHERE ФАМИЛИЯ==? AND ИМЯ==? AND ОТЧЕСТВО==? AND ГРУППА==?""", a).fetchall() == []:
                cursor.execute("""INSERT INTO 'СТУДЕНТ' VALUES (?,?,?,?)""", a)
                conn.commit()

            else:
                print('Такая запись уже существует')
                mb.showerror('Внимание', 'Студент уже есть в базе. Переход к созданию файла посещений')
        else:
            None

    mb.askokcancel('Загрузка групп', 'Загрузка завершена. Следующий шаг - создание файла посещений.')
    system('taskkill /IM EXCEL.EXE /F')
